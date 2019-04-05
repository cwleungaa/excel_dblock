from openpyxl import load_workbook
from openpyxl import Workbook
wb = load_workbook(filename = 'old.xlsx')

wb2 = Workbook()

ws = wb['Sheet1'] 
ws2 = wb2.active

for row in ws.rows:
	list_a = []
	list_b = []
	for cell in row:
		list_a.append(cell.value)
		if cell.value == None:
			continue
		else:
			list_b.append(cell.value)
	if list_b != []:		
		ws2.append(list_b)

	print(list_a)
	print(list_b)

wb2.save("new.xlsx")


		